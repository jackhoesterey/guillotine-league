from pathlib import Path
ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

F = "Arial"
BLUE  = Font(name=F, size=10, color="0000FF")           # inputs
BLK   = Font(name=F, size=10)                            # formulas / static
BLKB  = Font(name=F, size=10, bold=True)
GRN   = Font(name=F, size=10, color="008000")            # cross-sheet links
H1    = Font(name=F, size=16, bold=True, color="1F2430")
H2    = Font(name=F, size=11, bold=True, color="FFFFFF")
SMALL = Font(name=F, size=9, color="666666")
SMALLI= Font(name=F, size=9, color="666666", italic=True)

HDR   = PatternFill("solid", fgColor="1F2430")
YEL   = PatternFill("solid", fgColor="FFFF00")
BAND  = PatternFill("solid", fgColor="F2F4F7")
GRNF  = PatternFill("solid", fgColor="E6F4EA")
REDF  = PatternFill("solid", fgColor="FCE8E6")
AMBF  = PatternFill("solid", fgColor="FEF7E0")

thin = Side(style="thin", color="D0D4DC")
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def hdrrow(ws, row, labels, widths=None):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = H2; c.fill = HDR; c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 28

# ─────────────────────────────────────────────── REFERENCE
ref = wb.active
ref.title = "Reference"
ref["A1"] = "Reference tables — do not edit"; ref["A1"].font = H1
ref["A2"] = "Everything on the other tabs looks values up from here."; ref["A2"].font = SMALL

SEASON = [
 (1, 18, 0,   "none", "Survive",    1000),
 (2, 17, 14,  "none", "Survive",     975),
 (3, 16, 28,  "none", "Survive",     950),
 (4, 15, 42,  "none", "Survive",     900),
 (5, 14, 56,  "KC, CAR", "Navigate", 850),
 (6, 13, 70,  "CIN, DET, MIA, MIN", "Navigate", 820),
 (7, 12, 84,  "BUF, JAX, LAC, WSH", "Navigate", 780),
 (8, 11, 98,  "HOU, NO, NYG, SF",   "Navigate", 740),
 (9, 10, 112, "PIT, TEN",           "Navigate", 700),
 (10, 9, 126, "CHI, DEN, PHI, TB",  "Pivot",    650),
 (11, 8, 140, "ATL, CLE, GB, LAR, NE, SEA", "Pivot", 550),
 (12, 7, 154, "none (Thanksgiving)", "Pivot",   450),
 (13, 6, 168, "BAL, IND, LV, NYJ",  "Pivot",    350),
 (14, 5, 182, "ARI, DAL",           "Empty tank", 250),
 (15, 4, 196, "none", "Empty tank", 170),
 (16, 3, 210, "none", "Empty tank",  90),
 (17, 2, 224, "none", "Empty tank — champion crowned", 0),
]
hdrrow(ref, 4, ["Week","Teams alive","Chopped players in pool","Teams on bye","Stage","Target $ left at start of week"],
       [8,12,22,34,22,26])
for i,(w,t,p,b,s,tgt) in enumerate(SEASON):
    r = 5+i
    for col,val in enumerate([w,t,p,b,s,tgt], start=1):
        c = ref.cell(row=r, column=col, value=val); c.font = BLK; c.border = BOX
        c.alignment = Alignment(horizontal="center" if col!=4 and col!=5 else "left")
    ref.cell(row=r, column=6).number_format = '$#,##0'
    if i%2:
        for col in range(1,7): ref.cell(row=r,column=col).fill = BAND

BYES = {5:["KC","CAR"],6:["CIN","DET","MIA","MIN"],7:["BUF","JAX","LAC","WSH"],
        8:["HOU","NO","NYG","SF"],9:["PIT","TEN"],10:["CHI","DEN","PHI","TB"],
        11:["ATL","CLE","GB","LAR","NE","SEA"],13:["BAL","IND","LV","NYJ"],14:["ARI","DAL"]}
flat = sorted([(t,w) for w,ts in BYES.items() for t in ts])
hdrrow(ref, 24, ["NFL team","Bye week"], [12,12])
ref.column_dimensions["H"].width = 12; ref.column_dimensions["I"].width = 12
for i,(t,w) in enumerate(flat):
    r = 25+i
    for col,val in enumerate([t,w], start=1):
        c = ref.cell(row=r,column=col,value=val); c.font=BLK; c.border=BOX
        c.alignment = Alignment(horizontal="center")
ref["D24"] = "2026 bye weeks, all 32 teams."; ref["D24"].font = SMALL
ref["D25"] = "Source: NFL.com 2026 schedule release."; ref["D25"].font = SMALLI

TIERS = [
 ("Elite","Top-15 overall. Wins you weeks for the rest of the season.",250,200,175,9999),
 ("Real starter","Would start on your team every week from now on.",75,125,150,9999),
 ("Patch","Covers a bye or an injury for 2-4 weeks.",25,40,60,9999),
 ("Streamer","One good matchup, then droppable.",8,15,25,9999),
 ("Handcuff","Insurance body / lottery ticket.",3,5,10,15),
]
hdrrow(ref, 60, ["Tier","What it means","Max bid Wk 1-4","Max bid Wk 5-9","Max bid Wk 10-13","Max bid Wk 14-17"],
       [14,46,15,15,16,16])
for i,row in enumerate(TIERS):
    r = 61+i
    for col,val in enumerate(row, start=1):
        c = ref.cell(row=r,column=col,value=val); c.font=BLK; c.border=BOX
        if col>=3:
            c.number_format = '$#,##0'; c.alignment=Alignment(horizontal="center")
    if i%2:
        for col in range(1,7): ref.cell(row=r,column=col).fill = BAND
ref["A67"] = 'Wk 14-17 shows $9,999 to mean "no ceiling" — by then, money you do not spend is money you wasted.'
ref["A67"].font = SMALLI
ref["A68"] = "Source: bid ranges adapted from RotoWire and Fantasy Life guillotine FAAB analysis; see the playbook for the full reasoning."
ref["A68"].font = SMALLI
ref.sheet_view.showGridLines = False

# ─────────────────────────────────────────────── FAAB LOG
log = wb.create_sheet("FAAB Log")
log["A1"] = "FAAB Log"; log["A1"].font = H1
log["A2"] = "Log every bid you place — won or lost. Yellow cells are yours to fill in. Losing bids matter: they tell you what the market costs."
log["A2"].font = SMALL
hdrrow(log, 4, ["Week","Player","Pos","Tier","Bid $","Won? (Y/N)","Max bid allowed","Over ceiling?","$ left after","Notes"],
       [7,22,7,14,10,12,16,14,12,42])

LOGROWS = 120
tier_dv = DataValidation(type="list", formula1='"Elite,Real starter,Patch,Streamer,Handcuff"', allow_blank=True)
won_dv  = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
log.add_data_validation(tier_dv); log.add_data_validation(won_dv)

for i in range(LOGROWS):
    r = 5+i
    for col in range(1,7):
        c = log.cell(row=r, column=col); c.fill = YEL; c.font = BLUE; c.border = BOX
        c.alignment = Alignment(horizontal="center")
    log.cell(row=r, column=2).alignment = Alignment(horizontal="left")
    log.cell(row=r, column=5).number_format = '$#,##0'
    # max bid allowed = INDEX over tier table, column chosen by week stage
    log.cell(row=r, column=7, value=(
        f'=IF(OR($A{r}="",$D{r}=""),"",'
        f'INDEX(Reference!$C$61:$F$65,MATCH($D{r},Reference!$A$61:$A$65,0),'
        f'IF($A{r}<=4,1,IF($A{r}<=9,2,IF($A{r}<=13,3,4)))))'
    ))
    log.cell(row=r, column=8, value=(
        f'=IF(OR($E{r}="",$G{r}=""),"",IF($E{r}>$G{r},"OVER","ok"))'
    ))
    log.cell(row=r, column=9, value=(
        f'=IF($A{r}="","",Dashboard!$B$5-SUMIFS($E$5:$E{r},$F$5:$F{r},"Y"))'
    ))
    for col in (7,9):
        cc = log.cell(row=r, column=col); cc.number_format='$#,##0'; cc.font=BLK
        cc.border=BOX; cc.alignment=Alignment(horizontal="center")
    log.cell(row=r, column=8).font = BLK
    log.cell(row=r, column=8).border = BOX
    log.cell(row=r, column=8).alignment = Alignment(horizontal="center")
    log.cell(row=r, column=10).border = BOX
    log.cell(row=r, column=10).font = BLUE
    log.cell(row=r, column=10).fill = YEL
    tier_dv.add(log.cell(row=r, column=4))
    won_dv.add(log.cell(row=r, column=6))

# example row
ex = ["", "", "", "", "", "", "", "", "", ""]
log["A5"], log["B5"], log["C5"], log["D5"], log["E5"], log["F5"] = 2, "(example) J. Doe", "RB", "Patch", 21, "N"
log["J5"] = "Example row — overwrite it. Bid odd numbers ($21, not $20): ties lose players."
log["J5"].font = SMALLI
log.freeze_panes = "A5"
log.sheet_view.showGridLines = False

# ─────────────────────────────────────────────── WEEKLY RESULTS
wk = wb.create_sheet("Weekly Results")
wk["A1"] = "Weekly Results"; wk["A1"].font = H1
wk["A2"] = "Fill this in every Monday night. The 'chop line' is the score of the team that got eliminated — the lowest score in the league that week."
wk["A2"].font = SMALL
hdrrow(wk, 4, ["Week","My score","Chop line (lowest score)","League median","My margin over the chop","Safety","Team eliminated","Notes"],
       [7,11,22,15,22,14,20,38])
for i in range(17):
    r = 5+i; w = i+1
    c = wk.cell(row=r, column=1, value=w); c.font=BLK; c.border=BOX
    c.alignment=Alignment(horizontal="center")
    for col in (2,3,4,7,8):
        cc = wk.cell(row=r, column=col); cc.fill=YEL; cc.font=BLUE; cc.border=BOX
        cc.alignment=Alignment(horizontal="center" if col in (2,3,4) else "left")
        if col in (2,3,4): cc.number_format = '0.0'
    wk.cell(row=r, column=5, value=f'=IF(OR($B{r}="",$C{r}=""),"",$B{r}-$C{r})')
    wk.cell(row=r, column=6, value=(
        f'=IF($E{r}="","",IF($E{r}<=10,"DANGER",IF($E{r}<=25,"Close","Safe")))'
    ))
    for col in (5,6):
        cc = wk.cell(row=r, column=col); cc.font=BLK; cc.border=BOX
        cc.alignment=Alignment(horizontal="center")
    wk.cell(row=r, column=5).number_format = '0.0'
    if i%2:
        for col in range(1,9):
            if wk.cell(row=r,column=col).fill.fgColor.rgb != "00FFFF00":
                wk.cell(row=r,column=col).fill = BAND
wk["A23"] = "Margin under 10 points means you were one injury away from being eliminated. Two DANGER weeks in a row means spend FAAB now, whatever the budget curve says."
wk["A23"].font = SMALLI
wk.freeze_panes = "A5"
wk.sheet_view.showGridLines = False

# ─────────────────────────────────────────────── MY ROSTER
ros = wb.create_sheet("My Roster")
ros["A1"] = "My Roster"; ros["A1"].font = H1
ros["A2"] = "Type the NFL team code (DET, SEA, KC…) and the bye week fills itself in. The bye-week summary underneath shows where you are stacked up."
ros["A2"].font = SMALL
hdrrow(ros, 4, ["Slot","Player","Pos","NFL team","Bye week","Acquired","Cost $","Notes"],
       [10,22,7,11,11,16,10,40])
SLOTS = ["QB","RB1","RB2","WR1","WR2","TE","FLEX","BN1","BN2","BN3","BN4","BN5","BN6","BN7"]
for i,s in enumerate(SLOTS):
    r = 5+i
    c = ros.cell(row=r, column=1, value=s); c.font=BLKB; c.border=BOX
    c.alignment=Alignment(horizontal="center")
    for col in (2,3,4,6,7,8):
        cc = ros.cell(row=r,column=col); cc.fill=YEL; cc.font=BLUE; cc.border=BOX
        cc.alignment=Alignment(horizontal="center" if col in (3,4,7) else "left")
    ros.cell(row=r, column=7).number_format = '$#,##0'
    bc = ros.cell(row=r, column=5, value=(
        f'=IF($D{r}="","",IFERROR(INDEX(Reference!$B$25:$B$56,MATCH($D{r},Reference!$A$25:$A$56,0)),"?"))'
    ))
    bc.font=BLK; bc.border=BOX; bc.alignment=Alignment(horizontal="center")
    if i < 7:
        ros.cell(row=r, column=1).fill = GRNF

ros["A21"] = "Bye week exposure"; ros["A21"].font = Font(name=F, size=12, bold=True)
ros["A22"] = "How many of your 14 players are off that week. Anything over 4 in one week is a real problem — plan FAAB for it."
ros["A22"].font = SMALL
byeweeks = [5,6,7,8,9,10,11,13,14]
hdrrow(ros, 23, ["Week"]+[str(b) for b in byeweeks], None)
ros.cell(row=24, column=1, value="Players out").font = BLKB
ros.cell(row=24, column=1).border = BOX
for j,b in enumerate(byeweeks):
    c = ros.cell(row=24, column=2+j, value=f'=COUNTIFS($E$5:$E$18,{b})')
    c.font=BLK; c.border=BOX; c.alignment=Alignment(horizontal="center")
    h = ros.cell(row=23, column=2+j); h.alignment=Alignment(horizontal="center")
    ros.column_dimensions[get_column_letter(2+j)].width = 8
ros.column_dimensions["A"].width = 14
ros.sheet_view.showGridLines = False

# ─────────────────────────────────────────────── DASHBOARD
db = wb.create_sheet("Dashboard", 0)
db["A1"] = "Guillotine Command Centre — 2026"; db["A1"].font = H1
db["A2"] = "18 teams · Full PPR · $1,000 FAAB · one team chopped every week · champion crowned after Week 17"
db["A2"].font = SMALL
db["A3"] = "Fill in the two yellow cells below, then keep the FAAB Log and Weekly Results tabs up to date. Everything else calculates itself."
db["A3"].font = SMALLI

for col,w in zip("ABCDEFG",[31,21,3,24,15,4,46]): db.column_dimensions[col].width = w

def label(r, txt, bold=False):
    c = db.cell(row=r, column=1, value=txt); c.font = BLKB if bold else BLK; return c

db["A5"] = "Starting FAAB budget"; db["A5"].font = BLKB
db["B5"] = 1000; db["B5"].font = BLUE; db["B5"].fill = YEL
db["B5"].number_format = '$#,##0'; db["B5"].border = BOX
db["D5"] = "← type your league's budget"; db["D5"].font = SMALLI

db["A6"] = "Current NFL week"; db["A6"].font = BLKB
db["B6"] = 1; db["B6"].font = BLUE; db["B6"].fill = YEL
db["B6"].border = BOX; db["B6"].alignment = Alignment(horizontal="center")
db["D6"] = "← update this every Tuesday"; db["D6"].font = SMALLI

db["A8"] = "WHERE YOU STAND"; db["A8"].font = Font(name=F, size=12, bold=True, color="1F2430")

rows = [
 ("Teams still alive",            '=IFERROR(INDEX(Reference!$B$5:$B$21,MATCH($B$6,Reference!$A$5:$A$21,0)),"")', '0'),
 ("Chopped players in the pool",  '=IFERROR(INDEX(Reference!$C$5:$C$21,MATCH($B$6,Reference!$A$5:$A$21,0)),"")', '0'),
 ("Stage of the season",          '=IFERROR(INDEX(Reference!$E$5:$E$21,MATCH($B$6,Reference!$A$5:$A$21,0)),"")', '@'),
 ("Teams on bye this week",       '=IFERROR(INDEX(Reference!$D$5:$D$21,MATCH($B$6,Reference!$A$5:$A$21,0)),"")', '@'),
]
r = 9
for lab, f, fmt in rows:
    db.cell(row=r, column=1, value=lab).font = BLK
    c = db.cell(row=r, column=2, value=f); c.font = GRN; c.number_format = fmt
    c.border = BOX
    c.alignment = Alignment(horizontal="center", shrink_to_fit=(fmt == '@'))
    r += 1

db["A14"] = "THE MONEY"; db["A14"].font = Font(name=F, size=12, bold=True, color="1F2430")
money = [
 ("Spent so far",            "=SUMIFS('FAAB Log'!$E$5:$E$124,'FAAB Log'!$F$5:$F$124,\"Y\")", '$#,##0'),
 ("Budget remaining",        "=$B$5-$B$15", '$#,##0'),
 ("% of budget remaining",   "=IFERROR($B$16/$B$5,0)", '0.0%'),
 ("Target remaining, this week", '=IFERROR(INDEX(Reference!$F$5:$F$21,MATCH($B$6,Reference!$A$5:$A$21,0)),"")', '$#,##0'),
 ("Variance vs. target",     '=IF($B$18="","",$B$16-$B$18)', '$#,##0;($#,##0);-'),
]
r = 15
for lab, f, fmt in money:
    db.cell(row=r, column=1, value=lab).font = BLK
    c = db.cell(row=r, column=2, value=f); c.font = BLK; c.number_format = fmt
    c.border = BOX; c.alignment = Alignment(horizontal="center")
    r += 1
db["B16"].font = Font(name=F, size=12, bold=True)

db["A21"] = "Verdict"; db["A21"].font = BLKB
db["B21"] = (
 '=IF($B$19="","Enter the current week above.",'
 'IF(AND($B$6>=14,$B$16>150),"SPEND IT ALL. Money you carry past your final week is money you wasted.",'
 'IF($B$19<-150,"OVERSPENT. You are burning budget faster than the curve. Bid only on genuine starters from here.",'
 'IF($B$19>250,"HOARDING. The chop line rises every week and you are standing still. Start buying real starters now.",'
 '"On pace. Keep bidding to the tier ceilings on the Bid Guide."))))'
)
db["B21"].font = BLKB
db.merge_cells("B21:G21")
db["B21"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
db["B21"].fill = AMBF; db["B21"].border = BOX
db.row_dimensions[21].height = 32

db["A23"] = "HOW YOU ARE SCORING"; db["A23"].font = Font(name=F, size=12, bold=True, color="1F2430")
score = [
 ("Weeks survived",              "=COUNTIFS('Weekly Results'!$B$5:$B$21,\">0\")", '0'),
 ("Average margin over the chop","=IF(COUNT('Weekly Results'!$E$5:$E$21)=0,\"\",AVERAGE('Weekly Results'!$E$5:$E$21))", '0.0'),
 ("Closest call so far",         "=IF(COUNT('Weekly Results'!$E$5:$E$21)=0,\"\",MIN('Weekly Results'!$E$5:$E$21))", '0.0'),
 ("Danger weeks (margin under 10)","=COUNTIFS('Weekly Results'!$E$5:$E$21,\"<10\",'Weekly Results'!$E$5:$E$21,\">-9999\")", '0'),
]
r = 24
for lab, f, fmt in score:
    db.cell(row=r, column=1, value=lab).font = BLK
    c = db.cell(row=r, column=2, value=f); c.font = BLK; c.number_format = fmt
    c.border = BOX; c.alignment = Alignment(horizontal="center")
    r += 1

db["A29"] = "Sunday 11:00 AM ET — check every starter against the inactive list."
db["A29"].font = Font(name=F, size=11, bold=True, color="B00020")
db["A30"] = "More guillotine teams are eliminated by a player who did not play than by a bad draft. Set the alarm."
db["A30"].font = SMALLI

db["D8"] = "BID CEILINGS BY TIER"; db["D8"].font = Font(name=F, size=12, bold=True, color="1F2430")
db["D9"] = "For the current week, based on the week number above."
db["D9"].font = SMALL
hdr = ["Tier","Max bid now"]
for j,h in enumerate(hdr):
    c = db.cell(row=10, column=4+j, value=h); c.font=H2; c.fill=HDR; c.border=BOX
    c.alignment=Alignment(horizontal="center")
for i,t in enumerate(["Elite","Real starter","Patch","Streamer","Handcuff"]):
    r = 11+i
    c = db.cell(row=r, column=4, value=t); c.font=BLK; c.border=BOX
    c2 = db.cell(row=r, column=5, value=(
      f'=IFERROR(INDEX(Reference!$C$61:$F$65,MATCH($D{r},Reference!$A$61:$A$65,0),'
      f'IF($B$6<=4,1,IF($B$6<=9,2,IF($B$6<=13,3,4)))),"")'
    ))
    c2.font=BLK; c2.border=BOX; c2.number_format='$#,##0'
    c2.alignment=Alignment(horizontal="center")
db["D17"] = "Then adjust: if you would be bottom-five this week, bid the ceiling. If you are comfortably mid-pack, bid the low end or nothing at all."
db["D17"].font = SMALLI
db.merge_cells("D17:G18")
db["D17"].alignment = Alignment(wrap_text=True, vertical="top")

db["G5"] = "Blue text on yellow = you type it. Black = calculated. Green = pulled from the Reference tab."
db["G5"].font = SMALLI
db.merge_cells("G5:G6")
db["G5"].alignment = Alignment(wrap_text=True, vertical="top")

db.sheet_view.showGridLines = False

wb.save(str(ROOT / "Guillotine-FAAB-Tracker-2026.xlsx"))
print("saved")
